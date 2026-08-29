import json
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parents[1]
SOURCE_PATH = ROOT / "outputs" / "pending-known-issue-review" / "session-306-release-readiness.json"
REVIEW_PATH = ROOT / "outputs" / "pending-repair-first-review" / "opus-session-306-master-review-2026-08-29.json"
OUTPUT_PATH = ROOT / "outputs" / "pending-repair-first-review" / "Opus-306-Repair-First-Review-2026-08-29.xlsx"

source = json.loads(SOURCE_PATH.read_text(encoding="utf-8"))
review = json.loads(REVIEW_PATH.read_text(encoding="utf-8"))
issues = {row["id"]: row for row in source["issues"]}
decisions = {row["id"]: row for row in review["decisions"]}

wb = Workbook()
wb.remove(wb.active)

navy = "0B1324"
blue = "2563EB"
light_blue = "DBEAFE"
green = "DCFCE7"
amber = "FEF3C7"
red = "FEE2E2"
gray = "F3F4F6"
white = "FFFFFF"
thin = Side(style="thin", color="D7DCE5")

def style_sheet(ws, widths, freeze="A2"):
    ws.freeze_panes = freeze
    ws.auto_filter.ref = ws.dimensions
    ws.sheet_view.showGridLines = False
    for cell in ws[1]:
        cell.fill = PatternFill("solid", fgColor=navy)
        cell.font = Font(color=white, bold=True)
        cell.alignment = Alignment(vertical="center", wrap_text=True)
    ws.row_dimensions[1].height = 34
    for index, width in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(index)].width = width
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = Border(bottom=thin)

summary = wb.create_sheet("Summary")
summary.append(["OPUS 306 REPAIR-FIRST REVIEW", "VALUE"])
for key, value in [
    ("Status", "Held for user review — no deployment or production writes"),
    ("Rows reviewed", review["summary"]["rowsReviewed"]),
    ("Issues with approved commerce", review["summary"]["approvedCommerceIssues"]),
    ("Approved parts", review["summary"]["approvedParts"]),
    ("Approved buy links", review["summary"]["approvedBuyLinks"]),
    ("Verified parts", review["summary"]["verifiedParts"]),
    ("Verified buy links", review["summary"]["verifiedBuyLinks"]),
    ("No-commerce issues", review["summary"]["noCommerceIssues"]),
    ("Content corrections", review["summary"]["contentCorrections"]),
    ("Citation-gate ready", review["promotionReadiness"]["citationGateReady"]),
    ("Citation holds", len(review["promotionReadiness"]["citationHolds"])),
    ("Method", review["method"]),
]:
    summary.append([key, value])
style_sheet(summary, [34, 125])
summary.auto_filter.ref = summary.dimensions

held_ids = {row["id"] for row in review["promotionReadiness"]["citationHolds"]}

all_ws = wb.create_sheet("All 306 Issues")
all_headers = [
    "Make", "Model", "Years", "Title", "Lane", "Full How to Fix", "Decision",
    "Repair-first review", "Content correction", "Approved parts", "Fitment scope",
    "Prices", "Vendors", "Verified URLs", "Part verified", "Link verified", "Promotion gate", "Issue ID",
]
all_ws.append(all_headers)

approved_ws = wb.create_sheet("Approved Links")
approved_ws.append([
    "Make", "Model", "Years", "Known issue", "Part", "Fitment / branch scope", "Price",
    "Vendor", "Verified URL", "Part verified", "Link verified", "Issue ID",
])

correction_ws = wb.create_sheet("Corrections")
correction_ws.append(["Make", "Model", "Title", "Full How to Fix", "Required correction", "Issue ID"])

no_ws = wb.create_sheet("No Commerce")
no_ws.append(["Make", "Model", "Years", "Title", "Lane", "Full How to Fix", "Why no retail link", "Issue ID"])

hold_ws = wb.create_sheet("Citation Holds")
hold_ws.append(["Make", "Model", "Known issue", "Hold reason", "Issue ID"])

for issue in source["issues"]:
    decision = decisions[issue["id"]]
    parts = decision.get("fixParts", [])
    links = [link for part_row in parts for link in part_row.get("buyLinks", [])]
    years = ", ".join(str(year) for year in issue.get("years", []))
    all_ws.append([
        issue["make"], issue["model"], years, issue["title"], issue["lane"], issue["solution"],
        decision["disposition"], decision.get("repairFirst", ""), decision.get("contentCorrection", ""),
        "\n".join(part_row["component"] for part_row in parts),
        "\n".join(part_row["scope"] for part_row in parts),
        "\n".join(part_row.get("price", "") for part_row in parts),
        "\n".join(link["vendor"] for link in links),
        "\n".join(link["url"] for link in links),
        "TRUE" if parts and all(part_row.get("verified") is True for part_row in parts) else ("N/A" if not parts else "FALSE"),
        "TRUE" if links and all(link.get("verified") is True for link in links) else ("N/A" if not links else "FALSE"),
        "HOLD — DEAD SOURCES" if issue["id"] in held_ids else "READY",
        issue["id"],
    ])
    if parts:
        all_ws.cell(all_ws.max_row, 7).fill = PatternFill("solid", fgColor=green)
    else:
        all_ws.cell(all_ws.max_row, 7).fill = PatternFill("solid", fgColor=gray)
        no_ws.append([
            issue["make"], issue["model"], years, issue["title"], issue["lane"], issue["solution"],
            decision["disposition"], issue["id"],
        ])
    if decision.get("contentCorrection"):
        all_ws.cell(all_ws.max_row, 9).fill = PatternFill("solid", fgColor=red)
        correction_ws.append([
            issue["make"], issue["model"], issue["title"], issue["solution"],
            decision["contentCorrection"], issue["id"],
        ])
    if issue["id"] in held_ids:
        hold = next(row for row in review["promotionReadiness"]["citationHolds"] if row["id"] == issue["id"])
        hold_ws.append([issue["make"], issue["model"], issue["title"], hold["reason"], issue["id"]])
    for part_row in parts:
        for link in part_row.get("buyLinks", []):
            approved_ws.append([
                issue["make"], issue["model"], years, issue["title"], part_row["component"],
                part_row["scope"], part_row.get("price", ""), link["vendor"], link["url"],
                str(part_row.get("verified") is True).upper(), str(link.get("verified") is True).upper(), issue["id"],
            ])
            url_cell = approved_ws.cell(approved_ws.max_row, 9)
            url_cell.hyperlink = link["url"]
            url_cell.style = "Hyperlink"

style_sheet(all_ws, [16, 22, 19, 48, 18, 76, 42, 76, 70, 42, 74, 26, 20, 68, 14, 14, 22, 62])
style_sheet(approved_ws, [16, 22, 18, 48, 42, 72, 27, 20, 70, 14, 14, 62])
style_sheet(correction_ws, [16, 22, 48, 78, 78, 62])
style_sheet(no_ws, [16, 22, 18, 48, 18, 78, 48, 62])
style_sheet(hold_ws, [16, 22, 52, 60, 62])

for ws in (all_ws, approved_ws, correction_ws, no_ws, hold_ws):
    for row_idx in range(2, ws.max_row + 1):
        ws.row_dimensions[row_idx].height = 72

wb.save(OUTPUT_PATH)
print(json.dumps({
    "output": str(OUTPUT_PATH),
    "sheets": wb.sheetnames,
    "rows": {ws.title: ws.max_row - 1 for ws in wb.worksheets},
}, indent=2))
